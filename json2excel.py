import json
import openpyxl
from openpyxl import Workbook
import json
wb=Workbook()
worksheet=wb.active

def populateSheet(dictData,sheet):

    index=1
    for key, value in dictData.items():
        if key=="module name":
            sheet.cell(1,1,key)
            sheet.cell(1,2,value)
        if key=="timescale":
            sheet.cell(2,1,key)
            sheet.cell(2,2,value)
        if key=="allPortInfo":
            sheet.cell(index,1,"portName")
            sheet.cell(index,2,"direction")
            sheet.cell(index,3,"type")
            sheet.cell(index,4,"size")
            index_1=index+1
            for k,v in value.items():
                sheet.cell(index_1,1,k)
                sheet.cell(index_1,2,v["direction"])
                sheet.cell(index_1,3,v["type"])
                sheet.cell(index_1,4,v["size"])
                index_1+=1
        index+=1

with open ("portData") as jsonFile:
    dictData=json.load(jsonFile)

populateSheet(dictData,worksheet)
wb.save("portInfo.xlsx")
