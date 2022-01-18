import json
import openpyxl
from openpyxl import load_workbook

wb=load_workbook(filename="portInfo.xlsx")

def getSheetDict(sheet):
    dictData={}
    dictData["module name"]=sheet.cell(1,2).internal_value
    dictData["timescale"]=sheet.cell(2,2).internal_value
    row_index=4
    allPortInfo={}
    while(row_index):
        row_tuple=sheet[row_index]
        portName=row_tuple[0].internal_value
        if portName==None:
            row_start=0
            break
        details={}
        details["direction"]=row_tuple[1].internal_value
        details["type"]=row_tuple[2].internal_value
        details["size"]=row_tuple[3].internal_value
        allPortInfo[portName]=details
        row_index+=1
    dictData["allPortInfo"]=allPortInfo
    return dictData

for ws in wb:
    dictData=getSheetDict(ws)


#convert dict to json
data_json=json.dumps(dictData,indent=4)

#write file
with open("portDataNew","w") as f:
    f.write(data_json)
f.close()